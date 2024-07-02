import requests
from openpyxl import load_workbook
import logging
import sys
from datetime import datetime
# 配置日志输出格式
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Jira 服务器的基本信息
jira_url = 'https://www.Jira.com'
username = 'yabo.gong@outlook.com'
password = 'XXXXXX'
print("此程序读取名称为“获取JQL数量input.xlsx”，从第二行第二列开始读取JQL，并输出output文件")

while True:
    try:
        # 打开Excel文件
        logging.info('打开Excel文件...')
        workbook = load_workbook('获取JQL数量input.xlsx')

        # 选择要操作的工作表
        worksheet = workbook.active

        # 遍历每个单元格
        for row in worksheet.iter_rows(min_row=2, min_col=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for col, cell in enumerate(row, start=2):  # 从第2列开始
                # 检查单元格是否为空
                if cell.value is None:
                    continue

                # 获取JQL查询语句
                jql_query = cell.value

                # 构建 API 请求 URL
                api_url = jira_url + '/rest/api/latest/search'

                # 使用Jira REST API执行JQL查询并获取结果
                logging.info('执行JQL查询: %s', jql_query)
                response = requests.get(api_url, auth=(username, password), params={'jql': jql_query})

                # 解析API响应并获取结果
                if response.status_code == 200:
                    json_data = response.json()
                    total_issues = json_data['total']
                    logging.info('查询结果: %s 个问题', total_issues)

                    # 将结果填写到对应的单元格
                    cell.value = total_issues
                else:
                    logging.error('JQL查询失败：%s', response.text)

        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        # 保存Excel文件
        output_filename = f"获取JQL数量的output_{current_time}.xlsx"
        workbook.save(output_filename)
        logging.info('操作完成，结果保存在文件：%s', output_filename)

        # 关闭Excel文件
        workbook.close()
    except Exception as e:
        logging.error('发生错误：%s', str(e))

    # 提示是否重新运行代码
    user_input = input("是否重新运行代码？（按下回车键即可重新运行，输入任意其他内容退出）：")
    if user_input.strip() == "":
        continue
    else:
        break
sys.exit()

