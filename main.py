# -*- coding: utf-8 -*-
# Update Time ： 20230816
# Auth ： yabo.gong@outlook.com
import re
import sys
from jira import JIRA
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import urllib3
import time
from openpyxl import load_workbook
from multiprocessing import Process, freeze_support
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
def main():
    # 登录 JIRA
    def login_jira(username, password):
        options = {
            'server': 'https://www.jira.com/',  # 替换为你的 JIRA 实例的 URL
            'verify': False  # 如果你的 JIRA 实例使用自签名证书，请将 verify 设置为 False，否则改为 True
        }
        jira = JIRA(options, basic_auth=(username, password))
        return jira

    # 使用 JQL 查询获取问题列表
    def search_issues(jira, jql):
        issues = jira.search_issues(jql, maxResults=None)
        for i, issue in enumerate(issues):
            print(f"正在处理第{i+1}条数据...")
            print("关键词：", issue.key)
            # print("概要：", issue.fields.summary)
            # 处理数据...
        return issues

    # 处理问题数据并创建 DataFrame
    def process_data(issues):
        # 处理数据逻辑

        data = []
        st_data = []
        sw_data = []
        fae_data = []

        for issue in issues:
            if not issue.fields:
                continue

            created_date = datetime.datetime.strptime(issue.fields.created, "%Y-%m-%dT%H:%M:%S.%f%z")
            created_date = created_date.strftime("%Y-%m-%d")
            updated_date = datetime.datetime.strptime(issue.fields.updated, "%Y-%m-%dT%H:%M:%S.%f%z")
            updated_date = updated_date.strftime("%Y-%m-%d")

            components = ", ".join(
                [component.name for component in issue.fields.components]) if issue.fields.components else None

            status_name = issue.fields.status.name if issue.fields.status else None
            bug_severity = issue.fields.customfield_11044.value if issue.fields.customfield_11044 else None
            assignee_display_name = issue.fields.assignee.displayName if issue.fields.assignee else None

            data.append({
                '平台': issue.key.split('-')[0],
                '关键词': issue.key,
                '概要': issue.fields.summary,
                '状态': status_name,
                'BUG发现的项目': issue.fields.customfield_11043,
                'BUG严重等级': bug_severity,
                '模块': components,
                '优先级': issue.fields.priority.name if issue.fields.priority else None,
                'Customer客户名称': issue.fields.customfield_11029,
                '经办人': assignee_display_name,
                '报告人': issue.fields.reporter.displayName,
                '创建日期': created_date,
                '已更新': updated_date,
            })

            if status_name and status_name in st_status_list:
                st_data.append({
                    '平台': issue.key.split('-')[0],
                    '关键词': issue.key,
                    '概要': issue.fields.summary,
                    '状态': status_name,
                    'BUG发现的项目': issue.fields.customfield_11043,
                    '优先级': issue.fields.priority.name if issue.fields.priority else None,
                    'BUG严重等级': bug_severity,
                    '经办人': assignee_display_name,
                    '报告人': issue.fields.reporter.displayName,
                    '创建日期': created_date,
                    '已更新': updated_date,
                })

            if status_name and status_name in sw_status_list:
                sw_data.append({
                    '平台': issue.key.split('-')[0],
                    '关键词': issue.key,
                    '概要': issue.fields.summary,
                    '状态': status_name,
                    'BUG发现的项目': issue.fields.customfield_11043,
                    '优先级': issue.fields.priority.name if issue.fields.priority else None,
                    'BUG严重等级': bug_severity,
                    '经办人': assignee_display_name,
                    '报告人': issue.fields.reporter.displayName,
                    '创建日期': created_date,
                    '已更新': updated_date,
                })

            if status_name and status_name in fae_status_list:
                fae_data.append({
                    '平台': issue.key.split('-')[0],
                    '关键词': issue.key,
                    '概要': issue.fields.summary,
                    '状态': status_name,
                    '模块': components,
                    '优先级': issue.fields.priority.name if issue.fields.priority else None,
                    'Customer客户名称': issue.fields.customfield_11029,
                    '经办人': assignee_display_name,
                    '报告人': issue.fields.reporter.displayName,
                    '创建日期': created_date,
                    '已更新': updated_date,
                })

            if status_name and status_name in fae_status_list:
                fae_data.append({
                    '平台': issue.key.split('-')[0],
                    '关键词': issue.key,
                    '概要': issue.fields.summary,
                    '状态': status_name,
                    '模块': components,
                    '优先级': issue.fields.priority.name if issue.fields.priority else None,
                    'Customer客户名称': issue.fields.customfield_11029,
                    '经办人': assignee_display_name,
                    '报告人': issue.fields.reporter.displayName,
                    '创建日期': created_date,
                    '已更新': updated_date,
                })
        df = pd.DataFrame(data)
        st_df = pd.DataFrame(st_data)
        sw_df = pd.DataFrame(sw_data)
        fae_df = pd.DataFrame(fae_data)
        return output_file, df, st_df, sw_df, fae_df

    # 将 DataFrame 写入指定 sheet
    def write_to_excel(output_file, df, st_df, sw_df, fae_df):
        wb = Workbook()
        ws_all = wb.active
        ws_all.title = '数据源'
        ws_st = wb.create_sheet('ST未解决bug')
        ws_sw = wb.create_sheet('SW未解决bug')
        ws_fae = wb.create_sheet('FAE未解决bug')

        # 过滤非法字符
        if '概要' in df.columns:
            df['概要'] = df['概要'].apply(lambda x: re.sub(r'[\000-\010\013\014\016-\037]', '', x))
        if '概要' in st_df.columns:
            st_df['概要'] = st_df['概要'].apply(lambda x: re.sub(r'[\000-\010\013\014\016-\037]', '', x))
        if '概要' in sw_df.columns:
            sw_df['概要'] = sw_df['概要'].apply(lambda x: re.sub(r'[\000-\010\013\014\016-\037]', '', x))
        if '概要' in fae_df.columns:
            fae_df['概要'] = fae_df['概要'].apply(lambda x: re.sub(r'[\000-\010\013\014\016-\037]', '', x))

        # 按照创建日期倒序排序
        df = df.sort_values(by='创建日期', ascending=False)

        # 如果存在"创建日期"列，则进行排序和处理
        if '创建日期' in st_df.columns:
            st_df = st_df.sort_values(by='创建日期', ascending=False)
            st_df.insert(0, '序号', range(1, len(st_df) + 1))
            st_df['持续时长（天）'] = (datetime.datetime.now() - pd.to_datetime(st_df['创建日期'])).dt.days
            for row in dataframe_to_rows(st_df, index=False, header=True):
                ws_st.append(row)

        if '创建日期' in sw_df.columns:
            sw_df = sw_df.sort_values(by='创建日期', ascending=False)
            sw_df.insert(0, '序号', range(1, len(sw_df) + 1))
            sw_df['持续时长（天）'] = (datetime.datetime.now() - pd.to_datetime(sw_df['创建日期'])).dt.days
            for row in dataframe_to_rows(sw_df, index=False, header=True):
                ws_sw.append(row)

        if '创建日期' in fae_df.columns:
            fae_df = fae_df.sort_values(by='创建日期', ascending=False)
            fae_df.insert(0, '序号', range(1, len(fae_df) + 1))
            fae_df['持续时长（天）'] = (datetime.datetime.now() - pd.to_datetime(fae_df['创建日期'])).dt.days
            for row in dataframe_to_rows(fae_df, index=False, header=True):
                ws_fae.append(row)

        # 添加序号列和持续时长（天）列到数据源 sheet
        df.insert(0, '序号', range(1, len(df) + 1))
        df['持续时长（天）'] = (datetime.datetime.now() - pd.to_datetime(df['创建日期'])).dt.days
        for row in dataframe_to_rows(df, index=False, header=True):
            ws_all.append(row)

        now = datetime.datetime.now()
        current_date = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")  # 获取当前日期并格式化为"YYYY-MM-DD"的字符串
        output_file = f"{output_file}_{current_date}.xlsx"  # 将日期添加到输出文件名中
        wb.save(output_file)
        print("有问题请联系：yabo.gong@outlook.com")

    if __name__ == "__main__":
        while True:
            freeze_support()
            # ST和SW状态列表
            st_status_list = ['ST_Open', 'SW_Resolved', 'Need Info', 'ST_Pending','SW_RESOVLED_INTERNAL']  # 替换为你的 ST 状态列表
            sw_status_list = ['STL_Checked', 'SPM_Assigned', 'DO', 'PENDING', 'Working', '原厂分析', 'ST_Reopen','Initial','Reopen','WAIT 3rd PARTY','WAIT OFFICIAL RELEASE']  # 替换为你的 SW 状态列表
            fae_status_list = ['WAIT FAE INFO','Resolved', 'RELEASED', 'WORKED AROUND', 'Patched', 'HW AE', '未开始','已解决'  ]  # 替换为你的FAE 状态列表
            # 登录 JIRA
            username = 'yabo.gong@outlook.com'
            password = 'XXXXXX'
            jira = login_jira(username, password)
            print("成功登录JIRA。")
            try:
                # 加载Excel文件
                df = pd.read_excel('JQL数据查询input.xlsx', sheet_name=0)
                output_files = df['名称'].tolist()  # 将A列的数据转换为列表
                jql_data = df['JQL'].tolist()
                # 遍历每个输出文件名和对应的JQL数据
                for output_file, jql in zip(output_files, jql_data) :
                    # 使用 Jira 连接检索器中的 bug
                    issues = search_issues(jira, jql)

                    # 处理数据并创建 DataFrame
                    output_file, df, st_df, sw_df, fae_df = process_data(issues)

                    # 将数据写入 Excel 文件的多个 sheet 页
                    write_to_excel(output_file, df, st_df, sw_df, fae_df)  # 将output_file作为单独的参数传递
                    print(f"名称为: {output_file} 的数据已成功写入到 {output_file}.xlsx 文件。")
                break  # 退出循环
            except FileNotFoundError:
                # 如果未找到input.xlsx文件，则提示手动输入JQL
                output_file = [input("请输入输出文件名（可为空）: ")]
                jql_data = [input("请输入JQL（不能为空）: ")]

            # 使用 Jira 连接检索器中的 bug
            print("正在检索Jira数据，请稍候...")
            issues = search_issues(jira, jql_data[0])  # 只使用第一个JQL进行检索
            print(f"共找到{len(issues)}条数据。")

            # 处理数据并创建 DataFrame
            print("正在处理Jira数据...")
            output_file, df, st_df, sw_df, fae_df = process_data(issues)
            print("数据处理完成。")

            # 将数据写入 Excel 文件的多个 sheet 页
            print("正在将数据写入Excel文件...")
            write_to_excel(output_file, df, st_df, sw_df, fae_df)
if __name__ == "__main__":
    max_attempts = 6  # 最大尝试次数
    attempts = 0

    while attempts < max_attempts:
        try:
            main()  # 执行主要代码逻辑
        except Exception as e:
            print(f"Error occurred: {e}")
            attempts += 1
            print(f"重试... 次数 {attempts}/{max_attempts}")
            time.sleep(1)  # 休眠一段时间后再次尝试
        else:
            break  # 如果没有发生错误，则跳出循环
    else:
        print("输出的JQL有误，请核对后再次输入")